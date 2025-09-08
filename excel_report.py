import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import io
import base64
from tkinter import *
from tkinter import filedialog, messagebox
from tkinter.ttk import *
import os
from datetime import datetime

class ExcelReportGenerator:
    def __init__(self):
        self.root = Tk()
        self.root.title("Excel Report Generator")
        self.root.geometry("800x600")
        self.root.configure(bg='#f0f0f0')
        
        self.csv_file = None
        self.df = None
        
        self.setup_ui()
        
    def setup_ui(self):
        # Title
        title_label = Label(self.root, text="Excel Report Generator", 
                           font=('Arial', 20, 'bold'))
        title_label.pack(pady=20)
        
        # File selection frame
        file_frame = Frame(self.root)
        file_frame.pack(pady=10, padx=20, fill=X)
        
        self.file_label = Label(file_frame, text="No file selected", 
                               font=('Arial', 10))
        self.file_label.pack(side=LEFT, padx=10)
        
        select_btn = Button(file_frame, text="Select CSV File", 
                           command=self.select_file)
        select_btn.pack(side=RIGHT, padx=10)
        
        # Preview frame
        preview_frame = LabelFrame(self.root, text="Data Preview", padding=10)
        preview_frame.pack(pady=10, padx=20, fill=BOTH, expand=True)
        
        # Treeview for data preview
        self.tree = Treeview(preview_frame)
        self.tree.pack(fill=BOTH, expand=True)
        
        # Scrollbars for treeview
        v_scrollbar = Scrollbar(preview_frame, orient=VERTICAL, command=self.tree.yview)
        v_scrollbar.pack(side=RIGHT, fill=Y)
        self.tree.configure(yscrollcommand=v_scrollbar.set)
        
        h_scrollbar = Scrollbar(preview_frame, orient=HORIZONTAL, command=self.tree.xview)
        h_scrollbar.pack(side=BOTTOM, fill=X)
        self.tree.configure(xscrollcommand=h_scrollbar.set)
        
        # Options frame
        options_frame = LabelFrame(self.root, text="Report Options", padding=10)
        options_frame.pack(pady=10, padx=20, fill=X)
        
        self.include_charts = BooleanVar(value=True)
        charts_check = Checkbutton(options_frame, text="Include Charts", 
                                  variable=self.include_charts)
        charts_check.pack(side=LEFT, padx=10)
        
        self.include_summary = BooleanVar(value=True)
        summary_check = Checkbutton(options_frame, text="Include Summary Statistics", 
                                   variable=self.include_summary)
        summary_check.pack(side=LEFT, padx=10)
        
        self.include_pivot = BooleanVar(value=True)
        pivot_check = Checkbutton(options_frame, text="Include Pivot Tables", 
                                 variable=self.include_pivot)
        pivot_check.pack(side=LEFT, padx=10)
        
        # Generate button
        generate_btn = Button(self.root, text="Generate Excel Report", 
                             command=self.generate_report, style='Accent.TButton')
        generate_btn.pack(pady=20)
        
        # Status bar
        self.status_var = StringVar(value="Ready to generate reports")
        status_bar = Label(self.root, textvariable=self.status_var, 
                          relief=SUNKEN, anchor=W)
        status_bar.pack(side=BOTTOM, fill=X)
    
    def select_file(self):
        file_path = filedialog.askopenfilename(
            title="Select CSV File",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        
        if file_path:
            self.csv_file = file_path
            self.file_label.config(text=f"Selected: {os.path.basename(file_path)}")
            self.load_csv_preview()
    
    def load_csv_preview(self):
        try:
            self.df = pd.read_csv(self.csv_file)
            self.status_var.set(f"Loaded CSV with {len(self.df)} rows and {len(self.df.columns)} columns")
            
            # Clear existing data
            self.tree.delete(*self.tree.get_children())
            
            # Set up columns
            self.tree["columns"] = list(self.df.columns)
            self.tree["show"] = "headings"
            
            # Configure column headings
            for col in self.df.columns:
                self.tree.heading(col, text=col)
                self.tree.column(col, width=100, anchor=CENTER)
            
            # Insert data (first 100 rows for preview)
            for index, row in self.df.head(100).iterrows():
                self.tree.insert("", "end", values=list(row))
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load CSV file:\n{str(e)}")
            self.status_var.set("Error loading CSV file")
    
    def create_styled_worksheet(self, workbook, title):
        """Create a worksheet with consistent styling"""
        ws = workbook.create_sheet(title=title)
        
        # Define styles
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        data_font = Font(name='Arial', size=10)
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        return ws, header_font, header_fill, data_font, border
    
    def safe_adjust_column_widths(self, ws, num_columns, start_row=1, data_rows=None):
        """Safely adjust column widths without iterating over merged cells"""
        try:
            for col_num in range(1, num_columns + 1):
                max_length = 10  # Minimum width
                col_letter = get_column_letter(col_num)
                
                # Check specific cells instead of iterating over columns
                if data_rows:
                    # Check header
                    header_cell = ws.cell(row=start_row, column=col_num)
                    if header_cell.value:
                        max_length = max(max_length, len(str(header_cell.value)))
                    
                    # Check sample data rows (first 10 to avoid performance issues)
                    sample_rows = min(10, data_rows)
                    for row_offset in range(1, sample_rows + 1):
                        cell = ws.cell(row=start_row + row_offset, column=col_num)
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                
                # Set width with reasonable limits
                adjusted_width = min(max(max_length + 2, 12), 30)
                ws.column_dimensions[col_letter].width = adjusted_width
                
        except Exception as e:
            # Fallback: set standard width
            for col_num in range(1, num_columns + 1):
                col_letter = get_column_letter(col_num)
                ws.column_dimensions[col_letter].width = 15
    
    def add_summary_statistics(self, workbook):
        """Add summary statistics worksheet"""
        ws, header_font, header_fill, data_font, border = self.create_styled_worksheet(workbook, "Summary Statistics")
        
        # Get numeric columns only
        numeric_cols = self.df.select_dtypes(include=[np.number]).columns.tolist()
        
        if not numeric_cols:
            ws.cell(row=1, column=1, value="No numeric columns found for statistics")
            self.safe_adjust_column_widths(ws, 1)
            return
        
        # Create summary statistics
        summary_stats = self.df[numeric_cols].describe()
        
        # Add title
        title_cell = ws.cell(row=1, column=1, value="Summary Statistics")
        title_cell.font = Font(size=16, bold=True)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(numeric_cols) + 1)
        
        # Add headers
        row_start = 3
        ws.cell(row=row_start, column=1, value="Statistic").font = header_font
        ws.cell(row=row_start, column=1).fill = header_fill
        
        for i, col in enumerate(numeric_cols, 2):
            header_cell = ws.cell(row=row_start, column=i, value=col)
            header_cell.font = header_font
            header_cell.fill = header_fill
        
        # Add data
        for i, stat in enumerate(summary_stats.index, row_start + 1):
            ws.cell(row=i, column=1, value=stat).font = data_font
            for j, col in enumerate(numeric_cols, 2):
                value = summary_stats.loc[stat, col]
                data_cell = ws.cell(row=i, column=j, value=round(value, 2) if pd.notnull(value) else 'N/A')
                data_cell.font = data_font
        
        # Apply borders
        for row_num in range(row_start, row_start + len(summary_stats.index) + 1):
            for col_num in range(1, len(numeric_cols) + 2):
                ws.cell(row=row_num, column=col_num).border = border
        
        # Adjust column widths
        self.safe_adjust_column_widths(ws, len(numeric_cols) + 1, row_start, len(summary_stats.index))
    
    def add_pivot_tables(self, workbook):
        """Add pivot table analysis"""
        ws, header_font, header_fill, data_font, border = self.create_styled_worksheet(workbook, "Pivot Analysis")
        
        # Add title
        title_cell = ws.cell(row=1, column=1, value="Data Analysis Summary")
        title_cell.font = Font(size=16, bold=True)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
        
        current_row = 3
        
        # Column information section
        section_cell = ws.cell(row=current_row, column=1, value="Column Analysis")
        section_cell.font = Font(size=14, bold=True)
        current_row += 2
        
        # Headers
        headers = ["Column Name", "Data Type", "Unique Values", "Null Count"]
        for i, header in enumerate(headers, 1):
            header_cell = ws.cell(row=current_row, column=i, value=header)
            header_cell.font = header_font
            header_cell.fill = header_fill
        
        current_row += 1
        
        # Column analysis data
        for col in self.df.columns:
            ws.cell(row=current_row, column=1, value=col).font = data_font
            ws.cell(row=current_row, column=2, value=str(self.df[col].dtype)).font = data_font
            ws.cell(row=current_row, column=3, value=self.df[col].nunique()).font = data_font
            ws.cell(row=current_row, column=4, value=self.df[col].isnull().sum()).font = data_font
            current_row += 1
        
        # Apply borders
        end_row = current_row - 1
        for row_num in range(5, end_row + 1):
            for col_num in range(1, 5):
                ws.cell(row=row_num, column=col_num).border = border
        
        # Adjust column widths
        self.safe_adjust_column_widths(ws, 4, 5, len(self.df.columns))
    
    def add_charts(self, workbook):
        """Add charts to the workbook"""
        ws, header_font, header_fill, data_font, border = self.create_styled_worksheet(workbook, "Charts & Visualizations")
        
        # Add title
        title_cell = ws.cell(row=1, column=1, value="Data Visualizations")
        title_cell.font = Font(size=16, bold=True)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
        
        # Get numeric columns
        numeric_cols = self.df.select_dtypes(include=[np.number]).columns.tolist()
        
        if len(numeric_cols) >= 2:
            # Create a simple bar chart with first two numeric columns
            chart = BarChart()
            chart.title = f"Comparison: {numeric_cols[0]} vs {numeric_cols[1]}"
            chart.y_axis.title = "Values"
            chart.x_axis.title = "Records"
            
            # Add sample data for chart (first 10 rows)
            sample_data = self.df[numeric_cols[:2]].head(10)
            
            # Add data to worksheet for chart reference
            start_row = 5
            for i, col in enumerate(sample_data.columns):
                header_cell = ws.cell(row=start_row, column=i+1, value=col)
                header_cell.font = header_font
                header_cell.fill = header_fill
            
            for i, (idx, row) in enumerate(sample_data.iterrows()):
                for j, value in enumerate(row):
                    data_cell = ws.cell(row=start_row+i+1, column=j+1, value=value)
                    data_cell.font = data_font
            
            # Create chart reference
            data = Reference(ws, min_col=1, min_row=start_row, 
                           max_row=start_row+len(sample_data), 
                           max_col=len(numeric_cols[:2]))
            chart.add_data(data, titles_from_data=True)
            
            # Position chart
            ws.add_chart(chart, "E5")
            
            # Adjust column widths
            self.safe_adjust_column_widths(ws, len(numeric_cols[:2]), start_row, len(sample_data))
        else:
            info_cell = ws.cell(row=5, column=1, value="Not enough numeric columns for chart generation")
            info_cell.font = data_font
            self.safe_adjust_column_widths(ws, 1)
    
    def add_raw_data(self, workbook):
        """Add raw data worksheet"""
        ws, header_font, header_fill, data_font, border = self.create_styled_worksheet(workbook, "Raw Data")
        
        # Add title
        title_cell = ws.cell(row=1, column=1, value=f"Raw Data Export - {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        title_cell.font = Font(size=14, bold=True)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(self.df.columns))
        
        # Add headers
        start_row = 3
        for col_num, column_title in enumerate(self.df.columns, 1):
            header_cell = ws.cell(row=start_row, column=col_num, value=column_title)
            header_cell.font = header_font
            header_cell.fill = header_fill
            header_cell.alignment = Alignment(horizontal='center')
        
        # Add data
        for row_num, (index, row) in enumerate(self.df.iterrows(), start_row + 1):
            for col_num, value in enumerate(row, 1):
                data_cell = ws.cell(row=row_num, column=col_num, value=value)
                data_cell.font = data_font
                data_cell.alignment = Alignment(horizontal='left')
        
        # Apply borders to all data
        for row_num in range(start_row, start_row + len(self.df) + 1):
            for col_num in range(1, len(self.df.columns) + 1):
                ws.cell(row=row_num, column=col_num).border = border
        
        # Adjust column widths
        self.safe_adjust_column_widths(ws, len(self.df.columns), start_row, len(self.df))
    
    def generate_report(self):
        if self.df is None:
            messagebox.showerror("Error", "Please select a CSV file first!")
            return
        
        try:
            # Ask for save location
            save_path = filedialog.asksaveasfilename(
                title="Save Excel Report",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            
            if not save_path:
                return
            
            self.status_var.set("Generating Excel report...")
            self.root.update()
            
            # Create workbook
            workbook = Workbook()
            
            # Remove default sheet
            workbook.remove(workbook.active)
            
            # Add raw data
            self.add_raw_data(workbook)
            
            # Add summary statistics if selected
            if self.include_summary.get():
                self.add_summary_statistics(workbook)
            
            # Add pivot analysis if selected
            if self.include_pivot.get():
                self.add_pivot_tables(workbook)
            
            # Add charts if selected
            if self.include_charts.get():
                self.add_charts(workbook)
            
            # Save the workbook
            workbook.save(save_path)
            
            self.status_var.set(f"Report generated successfully: {os.path.basename(save_path)}")
            messagebox.showinfo("Success", f"Excel report generated successfully!\nSaved as: {save_path}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate report:\n{str(e)}")
            self.status_var.set("Error generating report")
    
    def run(self):
        self.root.mainloop()

# Sample data generator for testing
def create_sample_data():
    """Create sample CSV files for testing"""
    # Sales data
    sales_data = {
        'Date': pd.date_range('2024-01-01', periods=100, freq='D'),
        'Product': np.random.choice(['Laptop', 'Phone', 'Tablet', 'Monitor', 'Keyboard'], 100),
        'Category': np.random.choice(['Electronics', 'Accessories'], 100),
        'Quantity': np.random.randint(1, 20, 100),
        'Price': np.random.uniform(50, 1500, 100).round(2),
        'Sales_Rep': np.random.choice(['Alice', 'Bob', 'Charlie', 'Diana'], 100),
        'Region': np.random.choice(['North', 'South', 'East', 'West'], 100)
    }
    
    sales_df = pd.DataFrame(sales_data)
    sales_df['Total_Sales'] = sales_df['Quantity'] * sales_df['Price']
    sales_df.to_csv('sample_sales_data.csv', index=False)
    
    # Employee data
    employee_data = {
        'Employee_ID': range(1, 51),
        'Name': [f'Employee_{i}' for i in range(1, 51)],
        'Department': np.random.choice(['IT', 'Sales', 'Marketing', 'HR', 'Finance'], 50),
        'Age': np.random.randint(25, 60, 50),
        'Salary': np.random.uniform(40000, 120000, 50).round(2),
        'Experience_Years': np.random.randint(1, 20, 50),
        'Performance_Score': np.random.uniform(3.0, 5.0, 50).round(1)
    }
    
    employee_df = pd.DataFrame(employee_data)
    employee_df.to_csv('sample_employee_data.csv', index=False)
    
    print("Sample CSV files created:")
    print("- sample_sales_data.csv")
    print("- sample_employee_data.csv")

if __name__ == "__main__":
    # Create sample data files
    create_sample_data()
    
    # Launch the application
    app = ExcelReportGenerator()
    app.run()